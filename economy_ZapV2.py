"""
===============================================================================
economy_ZapV2.py  -  ZapGT-1 VERSION 2 (deeper producer) GEOTHERMAL DOUBLET DISTRICT-HEATING ECONOMIC MODEL
===============================================================================
Master feasibility model.  Run from the SAME FOLDER as the V3 production model
(in Spyder: click play here).  V3 is the unchanged "physics master"; this script
runs alongside it and is FED from it.  It integrates four engines:

    V3 reservoir/well model ...  production operating point + mirror-well injection
    esp_geothermal.py .........  ESP + injection-pump electrical power
    geothermal_HE.py ..........  PHE sizing, buried-pipeline heat loss, delivered
                                 heat, DH circulation pump  (revenue = delivered)
    doublet_decline.py ........  Gringarten-Sauty thermal breakthrough T_prod(t)

and adds the full economics: CAPEX, OPEX, debt financing, escalation, tax/
depreciation, and metrics (NPV, IRR, simple & discounted payback, LCOH, DSCR)
plus a one-at-a-time tornado.  Console report + Excel workbook + plots.

ALL unit costs are EDITABLE ASSUMPTIONS (generic Croatian/EU feasibility values,
flagged below).  Replace with quotes when available.
===============================================================================
"""
import os, sys, math, io, contextlib
from dataclasses import dataclass, field, asdict
import numpy as np
import plot_style  # uniform figure style (applied on import)

# --- per-well thermal / gas constants (for Alves wellhead-T + Jelic props) ---
T_BH = 86.7      # bottomhole flowing (mixed inflow) temperature, C
GWR_M3M3 = 1.0   # gas-water ratio (treated as CO2)
SUBBASIN = "Sava"
GEO_GRAD = 0.0365
SALINITY = 1320
WELL_NAME = "ZapGT-1-V2"


THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, THIS_DIR)

# ===========================================================================
# 1) CONFIG  -  every input lives here
# ===========================================================================
@dataclass
class Config:
    # ---- project / financing -------------------------------------------------
    project_life_yr:        int   = 30
    discount_rate:          float = 0.05      # [-] real discount rate (input)
    debt_ratio:             float = 0.50      # [-] 0=all equity, 1=all debt
    loan_interest:          float = 0.035     # [-] HBOR/commercial energy loan
    loan_tenor_yr:          int   = 17        # [yr]
    tax_rate:               float = 0.18      # [-] Croatian corporate tax
    inflation:              float = 0.03      # [-] general (real analysis -> 0)
    heat_price_escal:       float = 0.03      # [-] heat-price escalation (manual)
    elec_price_escal:       float = 0.03      # [-] electricity escalation (manual)
    opex_escal:             float = 0.03      # [-] opex escalation (manual)

    # ---- energy market -------------------------------------------------------
    heat_price_eur_MWhth:   float = 54.3      # [EUR/MWhth] heat selling price
    elec_price_eur_MWhe:    float = 135.0     # [EUR/MWhe]  electricity (uniform)
    capacity_price_eur_kW_yr: float = 45.0    # [EUR/kWth/yr] DH capacity payment (REVENUE)
    co2_price_eur_t:        float = 70.0      # [EUR/tCO2]
    co2_as_revenue:         bool  = False     # Y/N: count CO2 credit as revenue?
    gas_boiler_eff:         float = 0.85
    ng_emission_t_MWh:      float = 0.202     # tCO2/MWh_gas (LHV)

    # ---- operation -----------------------------------------------------------
    FLH:                    float = 2000.0    # [h/yr] annual full-load hours (heat)
    dh_scenario:            str   = "B"       # "A"=60C, "B"=40C, "C"=15C return

    # ---- geometry / distances (inputs) --------------------------------------
    city_distance_m:        float = 3000.0   # Zapresic: 5 km one-way (x2 supply+return)    # one-way well -> city; pipeline = 2x (supply+return)
    injection_distance_m:   float = 1400.0    # distance to (future) injection well     # well -> injection well (single brine line)
    grid_distance_m:        float = 500.0     # to nearest MV connection point
    dh_pipe_DN_m:           float = 0.3076    # DN300 (upsized: 125 L/s -> ~1.7 m/s, not 3.6)
    brine_pipe_DN_m:        float = 0.3076    # DN300 injection line (125 L/s)

    # ---- doublet (MANUAL inputs - user controls) ----------------------------
    doublet_avg_flow_ls:    float = 93.75     # [L/s] annual-AVERAGE circulation flow; drives the GS thermal decline only (93.75/(9/12)=125 L/s peak)
    operating_months_per_yr:int   = 9          # [months/yr] doublet ON (rest = OFF). peak/design flow = avg/(months/12) -> sizes reservoir IPR, ESP, pumps, PHE
    doublet_spacing_m:      float = 1400.0     # producer<->injector spacing (= injection distance)     # producer<->injector spacing (FEFLOW/own)
    doublet_method:         str   = "GS"
    barends_dispersivity_m: float = 5.0       # [m] Barends longitudinal dispersivity a_L (per layer); used only when doublet_method="GS+Barends"
    barends_underburden:    bool  = True       # Barends bleeding: True=cap+base (h_eff=2x), False=cap only; used only when doublet_method="GS+Barends"

    # ---- CAPEX unit costs (EDITABLE feasibility assumptions) -----------------
    well_cost_eur_per_m:    float = 3950.0    # turnkey EUR/m
    prod_well_cost_eur:     float = 8087625.0  # NEW deep producer (3950 EUR/m x 2047.5 m)
    inj_well_depth_m:       float = 0.0       # injector = EXISTING ZapGT-1 (already drilled)
    esp_eur_per_kW:         float = 1300.0    # ESP pump+motor+oversize tubing (surface kW)
    esp_cable_eur_per_m:    float = 45.0      # downhole power cable
    esp_install_eur:        float = 120000.0  # service-company install (lump)
    injpump_eur_per_kW:     float = 700.0     # injection pump
    injpump_install_eur:    float = 40000.0
    phe_eur_per_m2:         float = 700.0     # plate heat exchanger
    phe_capacity_factor:    float = 1.0       # (unused) PHE redundancy factor — CAPEX now uses the single sized PHE area only
    prod_pipe_ID_m:         float = 0.200    # production tubing ID (max to fit in 9-5/8"): 125 L/s -> ~3.3 m/s
    inj_pipe_ID_m:          float = 0.200    # injection tubing ID (DN200): 125 L/s lower dP
    circ_eur_per_kW:        float = 600.0    # DH circulation pump
    plant_prod_eur:         float = 50000.0  # energy-plant container @ production well
    plant_inj_eur:          float = 30000.0  # container @ injection well
    dh_pipe_eur_per_m:      float = 1200.0     # buried pre-insulated DN300 (per metre of trench)
    brine_pipe_eur_per_m:   float = 700.0     # buried brine line DN300
    transformer_eur_per_kVA:float = 50.0
    mv_line_eur_per_m:      float = 120.0
    grid_connection_fee_eur:float = 40000.0
    eng_pct:                float = 0.03      # engineering/design
    contingency_pct:        float = 0.10      # contingency

    # ---- OPEX (EDITABLE) -----------------------------------------------------
    personnel_eur_yr:       float = 30000.0   # 1 service person (Croatian mean gross + contrib.)
    grid_capacity_charge_eur_kW_yr: float = 10.0   # grid DEMAND charge (COST) on connected kW
    sm_pct_surface:         float = 0.015     # service&maint, % of surface mech/elec capex
    sm_pct_wells:           float = 0.020     # service&maint, % of well capex
    chemicals_eur_yr:       float = 10000.0   # scaling/corrosion inhibitor + cleaning
    insurance_pct:          float = 0.002     # of total capex
    misc_opex_eur_yr:       float = 10000.0   # telemetry, land, admin
    esp_replace_interval_yr:int   = 8         # ESP replacement (capex event)
    injpump_replace_interval_yr: int = 12

    # ---- concession (Croatian, inputs) --------------------------------------
    field_area_km2:         float = 10.0
    concession_fixed_eur_km2: float = 132.72
    concession_var_pct:     float = 0.03      # 3% of (heat price * delivered kWh)

    # ---- injectivity --------------------------------------------------------
    injectivity_multiplier: float = 1.0       # 1.0 = mirror (=productivity); de-rate if known


# ===========================================================================
# 2) V3 INTERFACE  -  pull production point + mirror-well injection from V3
#    Tries to import the V3 reservoir model; otherwise uses the values already
#    derived from V3 for the 15 L/s design point (documented constants).
# ===========================================================================
def get_v3_results(target_flow_ls=125.0, reinj_T_C=60.0, cfg=None):
    """ZapGT-1 (single-layer karst dolomite) quantities feeding the economics (from main_zapgt1)."""
    # DOUBLET porosity = 0.10 (connected vuggy/fracture flow), decoupled from the
    # 3.3% matrix porosity used inside the IPR/reservoir model (left untouched).
    layers = [dict(h=161.0, phi=0.10, k=9128.92, T0=86.7, name='dolomite (deep)')]
    res = dict(
        prod_flow_ls   = target_flow_ls,   # 125.0 L/s design
        wellhead_T_C   = 81.0,             # 71 C reservoir, minimal ascent loss (cool well)
        Pwf_bar        = 195.0,            # flowing BHP @ z_ref(1617.5) for 125 L/s
        static_bar     = 201.4,           # = P_res (single layer; NOT extrapolated)
        drawdown_bar   = 6.4,
        dynamic_level_m= 120.0,
        esp_depth_m    = 300.0,            # very shallow OK (huge k, tiny drawdown)
        esp_intake_bar = 25.1,
        inj_bhp_bar    = 167.6,            # symmetric: static + drawdown
        inj_overpres_bar = 7.7,
        inj_depth_m    = 2047.5,
        layers         = layers,
        source         = "ZapGT-1 V2 fallback constants (deeper producer, 125 L/s)")
    try:
        from main_zapv2 import build_commingled, CALIBRATED_CONFIG       # type: ignore
        from pvt import bar_to_Pa, Pa_to_bar, GRAVITY                     # type: ignore
        comm = build_commingled(CALIBRATED_CONFIG)
        res["inj_depth_m"] = comm.z_ref
        static = comm.layers[0].P_res/1e5         # single-layer static = P_res
        Pwf = Pa_to_bar(comm.Pwf_at_q(target_flow_ls/1000.0))
        res["Pwf_bar"] = round(Pwf,1)
        res["static_bar"] = round(static,2)
        res["drawdown_bar"] = round(static - Pwf,1)
        res["inj_bhp_bar"] = round(2*static - Pwf,1)
        res["inj_overpres_bar"] = round(static - Pwf,1)
        rho = getattr(comm,"rho_wb",983.0); Twh = res["wellhead_T_C"]
        Psat = (10**(8.07131-1730.63/(233.426+Twh)))*0.00133322
        zmin = comm.z_ref - (Pwf-(Psat+3.0))*1e5/(rho*GRAVITY)
        esp = max(200.0, round((zmin+100)/50.0)*50.0)
        res["esp_depth_m"] = float(esp)
        res["esp_intake_bar"] = round(Pwf - rho*GRAVITY*(comm.z_ref-esp)/1e5,1)
        res["dynamic_level_m"] = round(comm.z_ref - (Pwf-2.2)*1e5/(rho*GRAVITY),0)
        res["source"] = "ZapGT-1 V2 live import (main_zapv2)"
    except Exception as e:
        res["import_note"] = f"(live import not active: {e}); using documented ZapGT constants"
    return res


# ===========================================================================
# 3) ENGINEERING INTERFACE  -  ESP, injection pump, DHS, doublet
# ===========================================================================
SCEN_RETURN = {"A": 60.0, "B": 40.0, "C": 15.0}

def _silent(fn, *a, **k):
    with contextlib.redirect_stdout(io.StringIO()):
        return fn(*a, **k)

# --- surface-pipeline hydraulics (so flow changes are checked, not silently fixed) ---
def _mu_water(T_C):
    """Dynamic viscosity of water [Pa.s] (Vogel correlation, ~0-100 C)."""
    return 2.414e-5 * 10.0**(247.8 / ((T_C + 273.15) - 140.0))

def _pipe_dP(Q_m3s, D_m, L_m, rho, mu, eps_m=4.6e-5):
    """Darcy-Weisbach pressure drop with Colebrook friction for a round pipe."""
    if D_m <= 0 or Q_m3s <= 0:
        return dict(v_ms=0.0, Re=0.0, f=0.0, dP_bar=0.0, dP_Pa=0.0)
    A  = math.pi * D_m**2 / 4.0
    v  = Q_m3s / A
    Re = rho * v * D_m / max(mu, 1e-9)
    if Re < 2300.0:
        f = 64.0 / max(Re, 1e-9)
    else:
        f = 0.25 / (math.log10(eps_m/D_m/3.7 + 5.74/Re**0.9))**2
        for _ in range(50):
            fn = 1.0 / (-2.0*math.log10(eps_m/D_m/3.7 + 2.51/(Re*math.sqrt(f))))**2
            if abs(fn - f) < 1e-9:
                f = fn; break
            f = fn
    dP = f * (L_m / D_m) * 0.5 * rho * v**2
    return dict(v_ms=v, Re=Re, f=f, dP_bar=dP/1e5, dP_Pa=dP)

def _dn_for_velocity(Q_m3s, v_target=2.5):
    """Minimum inner diameter [m] to keep pipe velocity <= v_target."""
    if Q_m3s <= 0 or v_target <= 0:
        return 0.0
    return math.sqrt(4.0 * Q_m3s / (math.pi * v_target))

VEL_CAUTION_MS = 2.5    # comfort/erosion limit for liquid transmission lines
VEL_HIGH_MS    = 3.0    # hard caution (erosion / noise / excessive dP)
DPDL_CAUTION_PA_M = 150.0   # DH transmission pressure-gradient guideline (Pa per m)
DPDL_HIGH_PA_M    = 300.0   # clearly excessive gradient -> enlarge carrier pipe

def run_engineering(cfg, v3):
    reinj_T = SCEN_RETURN[cfg.dh_scenario]      # DH return ~ reinjection T (approx)
    import esp_geothermal_ZapV2 as esp

    # ---- production ESP (lift = dynamic level; duty from V3) -----------------
    esp.PROD_FLOW_LS = v3["prod_flow_ls"]
    esp.PROD_TEMP_C  = v3["wellhead_T_C"]
    esp.PROD_DEPTH_M = v3["dynamic_level_m"]    # pumping lift height
    esp.PROD_BACKPRESSURE_BAR = 3.2
    esp.PROD_PIPE_ID_M = cfg.prod_pipe_ID_m
    prod = _silent(esp.calc_well_pump, "Prod", esp.PROD_DEPTH_M, esp.PROD_FLOW_LS,
                   esp.PROD_TEMP_C, esp.PROD_PUMP_EFF, esp.PROD_MOTOR_EFF,
                   esp.PROD_CABLE_EFF, esp.PROD_VSD_EFF, esp.PROD_PIPE_ID_M,
                   esp.PROD_PIPE_ROUGHNESS_M, esp.PROD_BACKPRESSURE_BAR, mode="production")

    # ---- injection pump (mirror well; required BHP from V3) ------------------
    esp.INJ_FLOW_LS = v3["prod_flow_ls"]
    esp.INJ_TEMP_C  = reinj_T
    esp.INJ_DEPTH_M = v3["inj_depth_m"]
    esp.INJ_RESERVOIR_BAR = v3["inj_bhp_bar"]
    esp.INJ_WELLHEAD_BAR  = 0.0
    esp.INJ_PIPE_ID_M = cfg.inj_pipe_ID_m
    inj = _silent(esp.calc_well_pump, "Inj", esp.INJ_DEPTH_M, esp.INJ_FLOW_LS,
                  esp.INJ_TEMP_C, esp.INJ_PUMP_EFF, esp.INJ_MOTOR_EFF,
                  esp.INJ_CABLE_EFF, esp.INJ_VSD_EFF, esp.INJ_PIPE_ID_M,
                  esp.INJ_PIPE_ROUGHNESS_M, esp.INJ_WELLHEAD_BAR, mode="injection",
                  reservoir_pressure_bar=esp.INJ_RESERVOIR_BAR)

    # ---- DHS: PHE + buried pipeline loss + delivered heat + circulator -------
    import geothermal_HE_ZapV2 as dhs
    dhs.T_brine_in = v3["wellhead_T_C"]
    dhs.Q_brine    = v3["prod_flow_ls"]
    dhs.L_pipe     = cfg.city_distance_m
    dhs.D_inner    = cfg.dh_pipe_DN_m
    R = _silent(dhs.run_with_fallback, cfg.dh_scenario, SCEN_RETURN[cfg.dh_scenario],
                None, None, verbose=False) or {}
    # extract delivered heat / PHE area / circulator electrical power
    Q_HE = R.get("Q_HE", R.get("Q_W", None))
    if Q_HE is None:                      # reconstruct from brine cooling if absent
        rho = 970.0; cp = 4185.0
        Q_HE = v3["prod_flow_ls"]/1000.0 * rho * cp * (v3["wellhead_T_C"] - R.get("T_brine_out", reinj_T+4))
    loss_W = (R.get("Q_loss_sup", 0.0) or 0.0) + (R.get("Q_loss_ret", 0.0) or 0.0)
    delivered_W = max(Q_HE - loss_W, 0.0)
    phe = R.get("PHE", {})
    # Use the ACTUAL sized exchanger area (full plate selection, real U & LMTD) — this is
    # what gets installed and what the EUR/m2 CAPEX is based on. Fall back to the thermo-
    # dynamic required area, then a rough Q/(U*LMTD) estimate only if both are absent.
    phe_area = phe.get("A_total_m2") or phe.get("A_req_m2") or phe.get("area_m2") or phe.get("A_m2")
    if phe_area is None:
        U = 3500.0; LMTD = max(R.get("approach_hot", 5.0), 4.0)
        phe_area = Q_HE / (U * LMTD)
    # circulator electrical power from DHS pump duty
    dp = R.get("dp_pump_Pa", 0.0) or 0.0
    Qv = R.get("Q_vol", dhs.Q_brine/1000.0) or (dhs.Q_brine/1000.0)
    circ_hyd_W = dp * Qv
    circ_kW = circ_hyd_W / (0.75*0.95*0.97) / 1000.0
    brine_out = R.get("T_brine_out", reinj_T + 4.0)

    # ---- surface pipeline hydraulics (flow-driven velocity & pressure-drop check) ----
    # Brine reinjection line (cooled brine -> injection well): NOT modelled elsewhere,
    # so compute it here from the live flow + the user's DN. The DH carrier pipe
    # velocity / pipe pressure-drop are already produced by geothermal_HE (hyd_sup/ret);
    # we surface them and flag both lines if velocity exceeds safe limits.
    Q_brine_m3s = v3["prod_flow_ls"] / 1000.0
    rho_b = 985.0                                   # cooled brine ~ injection temperature
    mu_b  = 1.10 * _mu_water(brine_out)             # brine ~10% above pure water
    brine_hyd    = _pipe_dP(Q_brine_m3s, cfg.brine_pipe_DN_m, cfg.injection_distance_m, rho_b, mu_b)
    brine_DN_min = _dn_for_velocity(Q_brine_m3s, VEL_CAUTION_MS)

    hyd_sup = R.get("hyd_sup", {}) or {}
    hyd_ret = R.get("hyd_ret", {}) or {}
    dh_v    = hyd_sup.get("v", 0.0) or 0.0
    dh_dP_pipe_bar = ((hyd_sup.get("dp_total", 0.0) or 0.0) + (hyd_ret.get("dp_total", 0.0) or 0.0)) / 1e5
    Q_dh_m3s  = R.get("Q_vol", 0.0) or 0.0
    dh_DN_min = _dn_for_velocity(Q_dh_m3s, VEL_CAUTION_MS)

    warns = []
    def _flag(name, v, dn_min, D):
        if v >= VEL_HIGH_MS:
            warns.append(f"{name}: velocity {v:.2f} m/s exceeds {VEL_HIGH_MS:.1f} m/s "
                         f"(erosion / noise / high dP) - enlarge to >= DN{1000*dn_min:.0f} (now DN{1000*D:.0f}).")
        elif v >= VEL_CAUTION_MS:
            warns.append(f"{name}: velocity {v:.2f} m/s above the {VEL_CAUTION_MS:.1f} m/s comfort limit "
                         f"- consider >= DN{1000*dn_min:.0f} (now DN{1000*D:.0f}).")
    _flag("Brine reinjection line", brine_hyd["v_ms"], brine_DN_min, cfg.brine_pipe_DN_m)
    _flag("DH carrier pipe",        dh_v,             dh_DN_min,    cfg.dh_pipe_DN_m)
    def _flag_grad(name, dP_bar, L_total_m):
        if L_total_m <= 0:
            return
        g = dP_bar * 1e5 / L_total_m
        if g >= DPDL_HIGH_PA_M:
            warns.append(f"{name}: pressure gradient {g:.0f} Pa/m exceeds {DPDL_HIGH_PA_M:.0f} Pa/m "
                         f"(circulator head {dP_bar:.1f} bar over {L_total_m:.0f} m is high) - enlarge the carrier pipe.")
        elif g >= DPDL_CAUTION_PA_M:
            warns.append(f"{name}: pressure gradient {g:.0f} Pa/m above the {DPDL_CAUTION_PA_M:.0f} Pa/m guideline "
                         f"({dP_bar:.1f} bar over {L_total_m:.0f} m).")
    _flag_grad("DH carrier pipe", dh_dP_pipe_bar, 2.0 * cfg.city_distance_m)

    hydraulics = dict(
        brine_line=dict(D_m=cfg.brine_pipe_DN_m, L_m=cfg.injection_distance_m,
                        v_ms=brine_hyd["v_ms"], Re=brine_hyd["Re"],
                        dP_bar=brine_hyd["dP_bar"], DN_min_m=brine_DN_min),
        dh_line=dict(D_m=cfg.dh_pipe_DN_m, L_one_way_m=cfg.city_distance_m,
                     v_ms=dh_v, dP_pipe_bar=dh_dP_pipe_bar,
                     dP_circuit_bar=(dp/1e5), DN_min_m=dh_DN_min),
        warnings=warns)

    return dict(reinj_T=reinj_T,
                esp_kW=prod["surface_kW"], esp_TDH=prod["TDH_m"],
                inj_kW=(inj["surface_kW"] if inj.get("pump_needed", True) else 0.0),
                inj_TDH=inj.get("TDH_m", 0.0),
                Q_HE_kW=Q_HE/1000.0, delivered_kW=delivered_W/1000.0,
                pipe_loss_kW=loss_W/1000.0, phe_area_m2=phe_area,
                circ_kW=circ_kW, brine_out_C=brine_out,
                brine_line_v_ms=brine_hyd["v_ms"], brine_line_dP_bar=brine_hyd["dP_bar"],
                dh_line_v_ms=dh_v, dh_line_dP_pipe_bar=dh_dP_pipe_bar,
                hydraulics=hydraulics,
                _prod=prod, _inj=inj, _dhs=R)


# ===========================================================================
# 4) ENERGY (year 0) + doublet decline over life
# ===========================================================================
def energy_profile(cfg, v3, eng):
    import doublet_decline_ZapV2 as dd
    rho, cp = 970.0, 4.185                       # kg/m3, kJ/kgK
    dT0 = v3["wellhead_T_C"] - eng["brine_out_C"]
    installed_kWth = v3["prod_flow_ls"]/1000.0 * rho * cp * dT0   # brine-side gross
    delivered_kWth = eng["delivered_kW"]                          # DH-side, after losses
    # doublet decline
    t_yr, T_prod = dd.doublet_temperature_decline(
        cfg.doublet_avg_flow_ls, cfg.doublet_spacing_m, eng["reinj_T"],
        v3["layers"], years=cfg.project_life_yr,
        rho_w=v3.get("jelic",{}).get("brine",(1000.,4184.))[0],
        cp_w=v3.get("jelic",{}).get("brine",(1000.,4184.))[1],
        rock_rho=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[0],
        rock_cp=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[1],
        imp_K=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[2])
    dT_t = np.maximum(T_prod - eng["reinj_T"], 1e-6)
    decline = dT_t / dT_t[0]                      # fraction vs year 0
    # annual delivered MWh per project year (1..life)
    yrs = np.arange(1, cfg.project_life_yr + 1)
    decl_yr = np.interp(yrs, t_yr, decline)
    delivered_MWh = delivered_kWth * cfg.FLH / 1000.0 * decl_yr
    # --- optional Barends (2010) comparison view (economics stay on GS above) ---
    barends = None
    if getattr(cfg, "doublet_method", "GS") == "GS+Barends":
        _j = v3.get("jelic", {})
        barends = dd.decline_layered(
            cfg.doublet_avg_flow_ls, cfg.doublet_spacing_m, eng["reinj_T"],
            v3["layers"], years=cfg.project_life_yr,
            rho_w=_j.get("brine",(1000.,4184.))[0], cp_w=_j.get("brine",(1000.,4184.))[1],
            rock_rho=_j.get("rock",(2589.9,931.8,2.5))[0], rock_cp=_j.get("rock",(2589.9,931.8,2.5))[1],
            imp_K=_j.get("rock",(2589.9,931.8,2.5))[2],
            a_L=getattr(cfg, "barends_dispersivity_m", 5.0),
            include_underburden=getattr(cfg, "barends_underburden", True))
    return dict(installed_kWth=installed_kWth, delivered_kWth=delivered_kWth,
                delivered_MWh_y0=delivered_kWth*cfg.FLH/1000.0,
                delivered_MWh=delivered_MWh, decline=decl_yr,
                T_prod_end=float(T_prod[-1]), t_yr=t_yr, T_prod=T_prod, barends=barends)


# ===========================================================================
# 5) CAPEX
# ===========================================================================
def build_capex(cfg, v3, eng):
    total_kVA = (eng["esp_kW"] + eng["inj_kW"] + eng["circ_kW"]) / 0.9 * 1.25  # +margin
    items = {}
    items["Production well (turnkey)"]   = cfg.prod_well_cost_eur
    items["Injection well (turnkey)"]    = cfg.well_cost_eur_per_m * cfg.inj_well_depth_m
    items["ESP system (pump+cable+VSD+install)"] = (eng["esp_kW"]*cfg.esp_eur_per_kW
                                                    + cfg.esp_cable_eur_per_m*v3["esp_depth_m"]
                                                    + cfg.esp_install_eur)
    items["Injection pump system"]       = eng["inj_kW"]*cfg.injpump_eur_per_kW + cfg.injpump_install_eur
    items["PHE (single, sized for duty)"] = eng["phe_area_m2"] * cfg.phe_eur_per_m2
    items["DH circulator main+reserve"]  = 2 * max(eng["circ_kW"],1.0) * cfg.circ_eur_per_kW
    items["Energy plant/container @ prod"]= cfg.plant_prod_eur
    items["Container @ injection well"]  = cfg.plant_inj_eur
    items["DH pipeline (supply+return)"] = cfg.dh_pipe_eur_per_m * 2.0 * cfg.city_distance_m
    items["Brine reinjection pipeline"]  = cfg.brine_pipe_eur_per_m * cfg.injection_distance_m
    items["Grid connection (transformer+line+fee)"] = (total_kVA*cfg.transformer_eur_per_kVA
                                                       + cfg.mv_line_eur_per_m*cfg.grid_distance_m
                                                       + cfg.grid_connection_fee_eur)
    subtotal = sum(items.values())
    items["Engineering & design"]        = cfg.eng_pct * subtotal
    items["Contingency"]                 = cfg.contingency_pct * subtotal
    items["_TOTAL_"]                     = sum(items.values())
    items["_total_kVA_"]                 = total_kVA
    return items


# ===========================================================================
# 6) OPEX (year-1 basis) + scheduled replacements
# ===========================================================================
def build_opex(cfg, v3, eng, ener, capex):
    elec_MWh = (eng["esp_kW"] + eng["inj_kW"] + eng["circ_kW"]) * cfg.FLH / 1000.0
    surface_capex = (capex["ESP system (pump+cable+VSD+install)"]
                     + capex["Injection pump system"] + capex["PHE (single, sized for duty)"]
                     + capex["DH circulator main+reserve"]
                     + capex["Energy plant/container @ prod"] + capex["Container @ injection well"]
                     + capex["Grid connection (transformer+line+fee)"])
    well_capex = capex["Production well (turnkey)"] + capex["Injection well (turnkey)"]
    op = {}
    op["Electricity (ESP+inj+circ)"] = elec_MWh * cfg.elec_price_eur_MWhe
    op["Grid demand charge"]         = capex["_total_kVA_"]*0.9 * cfg.grid_capacity_charge_eur_kW_yr
    op["Personnel (1 FTE)"]          = cfg.personnel_eur_yr
    op["Service & maintenance"]      = cfg.sm_pct_surface*surface_capex + cfg.sm_pct_wells*well_capex
    op["Chemicals / inhibitor / cleaning"] = cfg.chemicals_eur_yr
    op["Insurance"]                  = cfg.insurance_pct * capex["_TOTAL_"]
    op["Misc (telemetry/land/admin)"]= cfg.misc_opex_eur_yr
    op["_elec_MWh_"] = elec_MWh
    return op


# ===========================================================================
# 7) CASHFLOW + 8) METRICS
# ===========================================================================
def npv(rate, cfs):
    return sum(cf/(1.0+rate)**t for t, cf in enumerate(cfs))

def irr(cfs):
    from scipy.optimize import brentq
    try:
        return brentq(lambda r: npv(r, cfs), -0.95, 5.0, xtol=1e-6)
    except Exception:
        return float("nan")

def run_cashflow(cfg, v3, eng, ener, capex, opex):
    N = cfg.project_life_yr
    CAPEX = capex["_TOTAL_"]
    # depreciation: wells/pipeline 30 yr, equipment 15 yr (straight line)
    dep_long = (capex["Production well (turnkey)"]+capex["Injection well (turnkey)"]
                + capex["DH pipeline (supply+return)"]+capex["Brine reinjection pipeline"]) / 30.0
    dep_short = (CAPEX - 30*dep_long if False else
                 (CAPEX - (capex["Production well (turnkey)"]+capex["Injection well (turnkey)"]
                           +capex["DH pipeline (supply+return)"]+capex["Brine reinjection pipeline"]))) / 15.0
    # debt
    debt = cfg.debt_ratio * CAPEX
    equity = CAPEX - debt
    # annuity loan payment
    if debt > 0 and cfg.loan_interest > 0:
        i, n = cfg.loan_interest, cfg.loan_tenor_yr
        ann = debt * i / (1 - (1+i)**(-n))
    else:
        ann = debt / cfg.loan_tenor_yr if debt > 0 else 0.0

    rows = []
    bal = debt
    cum_disc = 0.0; payback_yr = None
    fcf = [-CAPEX]              # year 0 (project free cash flow, unlevered)
    equity_cf = [-equity]      # year 0 (levered, equity)
    co2_t = ener["delivered_MWh"]/cfg.gas_boiler_eff * cfg.ng_emission_t_MWh
    for y in range(1, N+1):
        esc_h = (1+cfg.heat_price_escal)**(y-1)
        esc_e = (1+cfg.elec_price_escal)**(y-1)
        esc_o = (1+cfg.opex_escal)**(y-1)
        delivered = ener["delivered_MWh"][y-1]
        # revenue
        rev_energy   = delivered * cfg.heat_price_eur_MWhth * esc_h
        rev_capacity = ener["installed_kWth"] * cfg.capacity_price_eur_kW_yr   # fixed contracted
        rev_co2      = (co2_t[y-1]*cfg.co2_price_eur_t) if cfg.co2_as_revenue else 0.0
        revenue = rev_energy + rev_capacity + rev_co2
        # opex (+ concession variable on delivered heat)
        concession = (cfg.field_area_km2*cfg.concession_fixed_eur_km2
                      + cfg.concession_var_pct*rev_energy)
        opex_y = (sum(v for k,v in opex.items() if not k.startswith("_"))*esc_e_if(k_is_elec=False)*esc_o
                  if False else
                  opex["Electricity (ESP+inj+circ)"]*esc_e
                  + (sum(v for k,v in opex.items() if not k.startswith("_") and k!="Electricity (ESP+inj+circ)"))*esc_o
                  + concession)
        # scheduled replacements (capex events)
        repl = 0.0
        if y % cfg.esp_replace_interval_yr == 0 and y < N:
            repl += eng["esp_kW"]*cfg.esp_eur_per_kW + cfg.esp_cable_eur_per_m*v3["esp_depth_m"]
        if y % cfg.injpump_replace_interval_yr == 0 and y < N:
            repl += eng["inj_kW"]*cfg.injpump_eur_per_kW
        # debt service
        interest = bal*cfg.loan_interest if y <= cfg.loan_tenor_yr and bal>0 else 0.0
        principal = (ann - interest) if y <= cfg.loan_tenor_yr and bal>0 else 0.0
        principal = min(principal, bal)
        bal = max(bal - principal, 0.0)
        # depreciation
        dep = dep_long + (dep_short if y <= 15 else 0.0)
        ebitda = revenue - opex_y
        ebt = ebitda - dep - interest
        tax = max(ebt, 0.0)*cfg.tax_rate
        net_income = ebt - tax
        # cash flows
        proj_fcf = ebitda - tax - repl                          # unlevered (for project NPV/IRR @ discount)
        eq_cf    = ebitda - tax - interest - principal - repl   # levered equity cash flow
        dscr = ebitda/(interest+principal) if (interest+principal) > 0 else float("inf")
        fcf.append(proj_fcf); equity_cf.append(eq_cf)
        # discounted payback (project)
        disc = proj_fcf/(1+cfg.discount_rate)**y
        cum_disc += disc
        if payback_yr is None and (cum_disc - CAPEX) >= 0:
            payback_yr = y
        rows.append(dict(year=y, delivered_MWh=delivered, revenue=revenue,
                         rev_energy=rev_energy, rev_capacity=rev_capacity, rev_co2=rev_co2,
                         opex=opex_y, concession=concession, repl=repl,
                         interest=interest, principal=principal, debt_bal=bal,
                         dep=dep, tax=tax, ebitda=ebitda, proj_fcf=proj_fcf,
                         equity_cf=eq_cf, dscr=dscr))
    # metrics
    proj_npv = npv(cfg.discount_rate, fcf)
    proj_irr = irr(fcf)
    eq_npv   = npv(cfg.discount_rate, equity_cf)
    eq_irr   = irr(equity_cf) if equity > 0 else float("nan")
    # simple payback
    cum = -CAPEX; simple_pb = None
    for y in range(1, N+1):
        cum += fcf[y]
        if simple_pb is None and cum >= 0: simple_pb = y
    # LCOH = (disc capex + disc opex) / disc delivered heat
    disc_costs = CAPEX + sum((rows[y-1]["opex"]+rows[y-1]["repl"])/(1+cfg.discount_rate)**y for y in range(1,N+1))
    disc_heat  = sum(rows[y-1]["delivered_MWh"]/(1+cfg.discount_rate)**y for y in range(1,N+1))
    lcoh = disc_costs/disc_heat if disc_heat>0 else float("nan")
    dscr_vals = [r["dscr"] for r in rows if math.isfinite(r["dscr"])]
    return dict(rows=rows, fcf=fcf, equity_cf=equity_cf,
                proj_npv=proj_npv, proj_irr=proj_irr, eq_npv=eq_npv, eq_irr=eq_irr,
                simple_payback=simple_pb, disc_payback=payback_yr, lcoh=lcoh,
                dscr_min=min(dscr_vals) if dscr_vals else float("inf"),
                dscr_avg=float(np.mean(dscr_vals)) if dscr_vals else float("inf"),
                debt=debt, equity=equity, ann=ann)

def esc_e_if(k_is_elec):   # tiny helper kept for clarity
    return 1.0


# ===========================================================================
# 9) RUN + REPORT
# ===========================================================================
def run(cfg=None):
    cfg = cfg or Config()
    # ---- DERIVE peak/design flow from the annual-average flow + operating duty ----
    # The doublet circulates at PEAK flow for `operating_months_per_yr` and is off the
    # rest of the year, so   avg = peak * (months/12)   ->   peak = avg / (months/12).
    # The PEAK flow sizes the reservoir IPR (Pwf, drawdown, injection BHP, layer split),
    # the ESP (depth/kW), the injection pump and the PHE. The AVERAGE flow is used only
    # by the Gringarten-Sauty thermal decline (see energy_profile()), which needs the
    # true time-averaged throughput to place the cold front correctly.
    _duty = max(cfg.operating_months_per_yr / 12.0, 1e-6)
    peak_flow_ls = cfg.doublet_avg_flow_ls / _duty
    v3  = get_v3_results(target_flow_ls=peak_flow_ls, reinj_T_C=SCEN_RETURN[cfg.dh_scenario], cfg=cfg)
    v3['avg_flow_ls']             = cfg.doublet_avg_flow_ls
    v3['operating_months_per_yr'] = cfg.operating_months_per_yr
    v3['duty_fraction']           = _duty
    v3['peak_flow_ls']            = peak_flow_ls   # == prod_flow_ls (design point)
    import wellbore_T, geothermo_props as gtp
    _rk = gtp.rock_props(v3['inj_depth_m'], SUBBASIN)
    _br = gtp.brine_props(T_BH, SALINITY)
    v3['T_bh_C'] = T_BH
    v3['wellhead_T_C'] = round(wellbore_T.wellhead_temperature(
        L_m=v3['inj_depth_m'], q_ls=v3['prod_flow_ls'], T_bh_C=T_BH,
        geo_grad_K_m=GEO_GRAD, k_e=_rk[2], GWR_m3m3=GWR_M3M3,
        dyn_level_m=v3.get('dynamic_level_m',0.0),
        P_bh_bar=v3['Pwf_bar'], P_wh_bar=3.2), 1)
    v3['jelic'] = dict(rock=_rk, brine=_br)
    eng = run_engineering(cfg, v3)
    ener= energy_profile(cfg, v3, eng)
    cap = build_capex(cfg, v3, eng)
    op  = build_opex(cfg, v3, eng, ener, cap)
    cf  = run_cashflow(cfg, v3, eng, ener, cap, op)
    return dict(cfg=cfg, v3=v3, eng=eng, ener=ener, capex=cap, opex=op, cf=cf)


def print_report(R):
    cfg, v3, eng, ener, cap, op, cf = (R["cfg"], R["v3"], R["eng"], R["ener"],
                                       R["capex"], R["opex"], R["cf"])
    L = "="*70
    print(L); print(" ZapGT-1 V2 (DEEPER PRODUCER) GEOTHERMAL DOUBLET DH  -  ECONOMIC FEASIBILITY"); print(L)
    print(f" V3 source: {v3['source']}")
    print(f" Production: {v3['prod_flow_ls']:.1f} L/s, wellhead {v3['wellhead_T_C']:.0f} C, "
          f"drawdown {v3['drawdown_bar']:.0f} bar, ESP @ {v3['esp_depth_m']:.0f} m (intake {v3['esp_intake_bar']:.0f} bar)")
    print(f" Injection (mirror): BHP {v3['inj_bhp_bar']:.0f} bar -> pump {eng['inj_kW']:.0f} kW")
    print(f" DH scenario {cfg.dh_scenario} (return {SCEN_RETURN[cfg.dh_scenario]:.0f} C); brine cooled to {eng['brine_out_C']:.1f} C")
    print("-"*70)
    print(f" Installed thermal capacity : {ener['installed_kWth']/1000:8.3f} MWth (brine side)")
    print(f" Delivered heat (post-loss) : {ener['delivered_kWth']/1000:8.3f} MWth ; {ener['delivered_MWh_y0']:8.0f} MWh/yr (yr1)")
    print(f" Pipeline heat loss         : {eng['pipe_loss_kW']:8.1f} kW ; PHE area {eng['phe_area_m2']:.0f} m2")
    print(f" Pumps: ESP {eng['esp_kW']:.0f} kW, injection {eng['inj_kW']:.0f} kW, circulator {eng['circ_kW']:.1f} kW ; elec {op['_elec_MWh_']:.0f} MWh/yr")
    print(f" Doublet: T_prod {v3['wellhead_T_C']:.0f} C -> {ener['T_prod_end']:.1f} C at yr {cfg.project_life_yr} "
          f"(spacing {cfg.doublet_spacing_m:.0f} m, avg {cfg.doublet_avg_flow_ls:.1f} L/s)")
    print("-"*70); print(" CAPEX breakdown [EUR]")
    for k, v in cap.items():
        if not k.startswith("_"): print(f"   {k:<42s} {v:>14,.0f}")
    print(f"   {'TOTAL CAPEX':<42s} {cap['_TOTAL_']:>14,.0f}")
    print("-"*70); print(" OPEX (year 1) [EUR/yr]")
    for k, v in op.items():
        if not k.startswith("_"): print(f"   {k:<42s} {v:>14,.0f}")
    print(f"   {'Concession (yr1)':<42s} {cf['rows'][0]['concession']:>14,.0f}")
    print("-"*70); print(" REVENUE (year 1) [EUR/yr]")
    r0 = cf["rows"][0]
    print(f"   {'Energy (delivered heat)':<42s} {r0['rev_energy']:>14,.0f}")
    print(f"   {'Capacity payment':<42s} {r0['rev_capacity']:>14,.0f}")
    print(f"   {'CO2 credit'+('' if cfg.co2_as_revenue else ' (informational)'):<42s} {r0['rev_co2']:>14,.0f}")
    print("-"*70); print(" FINANCING & METRICS")
    print(f"   Debt {cfg.debt_ratio*100:.0f}% = {cf['debt']:,.0f} EUR @ {cfg.loan_interest*100:.1f}% / {cfg.loan_tenor_yr} yr ; equity {cf['equity']:,.0f}")
    print(f"   Discount rate              : {cfg.discount_rate*100:.1f}%")
    print(f"   Project NPV                : {cf['proj_npv']:>14,.0f} EUR")
    print(f"   Project IRR                : {cf['proj_irr']*100:>13.1f} %")
    if cf['equity'] > 0:
        print(f"   Equity NPV / IRR           : {cf['eq_npv']:,.0f} EUR / {cf['eq_irr']*100:.1f} %")
    print(f"   Simple / discounted payback: {cf['simple_payback']} / {cf['disc_payback']} yr")
    print(f"   LCOH                       : {cf['lcoh']:>13.2f} EUR/MWhth   (vs price {cfg.heat_price_eur_MWhth:.0f})")
    print(f"   DSCR min / avg             : {cf['dscr_min']:.2f} / {cf['dscr_avg']:.2f}")
    print(L)


def tornado(R, save_dir):
    """One-at-a-time +/-20% sensitivity of project NPV."""
    base = R["cf"]["proj_npv"]
    levers = {
        "Heat price":        ("heat_price_eur_MWhth", 0.20),
        "Electricity price": ("elec_price_eur_MWhe", 0.20),
        "FLH":               ("FLH", 0.20),
        "CAPEX (well cost/m)":("well_cost_eur_per_m", 0.20),
        "Discount rate":     ("discount_rate", 0.20),
        "Doublet spacing":   ("doublet_spacing_m", 0.20),
        "Capacity price":    ("capacity_price_eur_kW_yr", 0.20),
    }
    res = []
    for name, (attr, pct) in levers.items():
        lo = Config(**{**asdict(R["cfg"]), attr: getattr(R["cfg"], attr)*(1-pct)})
        hi = Config(**{**asdict(R["cfg"]), attr: getattr(R["cfg"], attr)*(1+pct)})
        try:
            nlo = run(lo)["cf"]["proj_npv"]; nhi = run(hi)["cf"]["proj_npv"]
        except Exception:
            nlo = nhi = base
        res.append((name, nlo, nhi))
    try:
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        res.sort(key=lambda x: abs(x[2]-x[1]))
        names = [r[0] for r in res]; los = [r[1] for r in res]; his = [r[2] for r in res]
        fig, ax = plt.subplots(figsize=(9,5), dpi=140)
        y = np.arange(len(names))
        for i,(n,lo,hi) in enumerate(res):
            ax.barh(i, hi-lo, left=min(lo,hi), color="#2E75B6", alpha=0.8)
        ax.axvline(base, color="k", ls="--", lw=1, label=f"base NPV {base/1e6:.2f} M€")
        ax.set_yticks(y); ax.set_yticklabels(names); ax.set_xlabel("Project NPV [EUR]")
        ax.set_title("Tornado - project NPV sensitivity (+/-20%)"); ax.legend()
        fig.tight_layout(); p=os.path.join(save_dir,"economy_tornado.png"); fig.savefig(p); plt.close(fig)
        return p
    except Exception as e:
        return f"(tornado plot skipped: {e})"


def plots(R, save_dir):
    import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
    cf = R["cf"]; ener = R["ener"]; cfg = R["cfg"]
    out = []
    # cumulative discounted cash flow (payback)
    fig, ax = plt.subplots(figsize=(8,4.5), dpi=140)
    cum = np.cumsum([cf["fcf"][0]] + [cf["fcf"][y]/(1+cfg.discount_rate)**y for y in range(1,cfg.project_life_yr+1)])
    ax.plot(range(0,cfg.project_life_yr+1), np.array(cum)/1e6, marker="o", ms=3, color="#1F3864")
    ax.axhline(0, color="k", lw=0.8); ax.set_xlabel("Year"); ax.set_ylabel("Cumulative discounted cash flow [M€]")
    ax.set_title("Discounted payback"); ax.grid(alpha=0.3)
    fig.tight_layout(); p=os.path.join(save_dir,"economy_payback.png"); fig.savefig(p); plt.close(fig); out.append(p)
    # temperature decline
    from matplotlib.ticker import FormatStrFormatter
    fig, ax = plt.subplots(figsize=(8,5.8), dpi=140)
    ax.plot(ener["t_yr"], ener["T_prod"], color="#C00000")
    ax.set_xlabel("Year"); ax.set_ylabel("Producer temperature [\u00b0C]")
    ax.set_title(f"Gringarten-Sauty doublet decline (spacing {cfg.doublet_spacing_m:.0f} m, {cfg.doublet_avg_flow_ls:.0f} L/s avg)")
    _Tp = np.asarray(ener["T_prod"]); _lo, _hi = float(_Tp.min()), float(_Tp.max())
    if _hi - _lo < 10.0:
        _mid = 0.5*(_lo+_hi); _lo, _hi = _mid-6.0, _mid+6.0
    ax.set_ylim(_lo-1.0, _hi+1.0)
    ax.ticklabel_format(style="plain", axis="y", useOffset=False)
    ax.yaxis.set_major_formatter(FormatStrFormatter("%.1f"))
    ax.grid(alpha=0.3); fig.tight_layout()
    p=os.path.join(save_dir,"economy_Tdecline.png"); fig.savefig(p); plt.close(fig); out.append(p)
    return out


def to_excel(R, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as e:
        return f"(openpyxl missing: {e})"
    cfg, cap, op, cf, ener, eng, v3 = (R["cfg"],R["capex"],R["opex"],R["cf"],R["ener"],R["eng"],R["v3"])
    wb = Workbook(); hd = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F3864")
    def sheet(title, rows):
        ws = wb.create_sheet(title)
        for r,row in enumerate(rows,1):
            for c,val in enumerate(row,1):
                cell = ws.cell(r,c,val)
                if r==1: cell.font=hd; cell.fill=fill
        return ws
    wb.remove(wb.active)
    sheet("Summary", [["Metric","Value","Unit"],
        ["Production flow", v3["prod_flow_ls"], "L/s"],
        ["Installed capacity", round(ener["installed_kWth"]/1000,3), "MWth"],
        ["Delivered heat yr1", round(ener["delivered_MWh_y0"]), "MWh/yr"],
        ["Total CAPEX", round(cap["_TOTAL_"]), "EUR"],
        ["Project NPV", round(cf["proj_npv"]), "EUR"],
        ["Project IRR", round(cf["proj_irr"]*100,1), "%"],
        ["Discounted payback", cf["disc_payback"], "yr"],
        ["LCOH", round(cf["lcoh"],2), "EUR/MWhth"],
        ["DSCR min", round(cf["dscr_min"],2), "-"]])
    sheet("CAPEX", [["Item","EUR"]] + [[k,round(v)] for k,v in cap.items() if not k.startswith("_")]
                   + [["TOTAL", round(cap["_TOTAL_"])]])
    sheet("OPEX_yr1", [["Item","EUR/yr"]] + [[k,round(v)] for k,v in op.items() if not k.startswith("_")])
    cfh = ["year","delivered_MWh","revenue","opex","concession","repl","interest","principal",
           "debt_bal","dep","tax","ebitda","proj_fcf","equity_cf","dscr"]
    sheet("Cashflow", [cfh] + [[round(r[k],1) if isinstance(r[k],float) else r[k] for k in cfh] for r in cf["rows"]])
    wb.save(path); return path


if __name__ == "__main__":
    R = run()
    print_report(R)
    save_dir = os.environ.get("ECON_OUT", THIS_DIR)
    xp = to_excel(R, os.path.join(save_dir, "ZapV2_economy.xlsx"))
    pp = plots(R, save_dir)
    tp = tornado(R, save_dir)
    try:
        import doublet_viz
        _mp = max(R['v3']['layers'], key=lambda L: L['k'])
        vz = doublet_viz.visualize_doublet(
            dict(h=_mp['h'], k=_mp['k'], name=_mp.get('name','reservoir')),
            R['cfg'].doublet_avg_flow_ls, R['cfg'].doublet_spacing_m,
            R['eng']['reinj_T'], _mp['T0'], phi_doublet=_mp['phi'],
            well_name=WELL_NAME, save_dir=save_dir)
        print(' Viz   :', vz)
    except Exception as e:
        print(' Viz skipped:', e)
    print(f"\n Excel : {xp}\n Plots : {pp}\n Tornado: {tp}")
